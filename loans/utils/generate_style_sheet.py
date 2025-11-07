from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import LineChart, Reference
from decimal import Decimal

def export_amortization_excel(loan, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Amortization Schedule"

    # ----------- Loan Summary Section -----------
    summary_data = [
        ("Loan Amount", loan.amount),
        ("Term of Loan in Years", loan.term_years),
        ("Annual Interest Rate", f"{loan.annual_interest_rate}%"),
        ("Compound Periods per Year", 12),
        ("Payments per Year", 12),
        ("Monthly Payment", loan.monthly_payment),
        ("Number of Payments", loan.term_years * 12),
        ("Rate Per Period", f"{(loan.annual_interest_rate/Decimal(12)):.3f}%"),
        ("Total Payment", sum([r["total_payment"] for r in schedule])),
        ("Total Interest", sum([r["interest_component"] for r in schedule])),
    ]

    ws.append(["Loan Summary"])
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    for key, value in summary_data:
        ws[f"A{row}"] = key
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    start_table_row = row + 2

    # ----------- Table Header -----------
    headers = [
        "Year", "Cumulative Interest", "Cumulative Principal",
        "Balance", "Cumulative Payments", "Yearly Principal", "Yearly Interest"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_table_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ----------- Table Body -----------
    cumulative_interest = Decimal("0.00")
    cumulative_principal = Decimal("0.00")
    cumulative_payments = Decimal("0.00")

    table_row = start_table_row + 1

    for year in range(1, loan.term_years + 1):
        year_rows = schedule[(year - 1) * 12 : year * 12]

        yearly_interest = sum([r["interest_component"] for r in year_rows], Decimal("0.00"))
        yearly_principal = sum([r["principal_component"] for r in year_rows], Decimal("0.00"))
        yearly_payment = sum([r["total_payment"] for r in year_rows], Decimal("0.00"))

        cumulative_interest += yearly_interest
        cumulative_principal += yearly_principal
        cumulative_payments += yearly_payment
        balance = loan.amount - cumulative_principal

        row_data = [
            year, float(cumulative_interest), float(cumulative_principal),
            float(balance), float(cumulative_payments),
            float(yearly_principal), float(yearly_interest)
        ]

        for col, val in enumerate(row_data, start=1):
            ws.cell(row=table_row, column=col, value=val)

        table_row += 1

    # ----------- Chart (3 Lines) -----------
    chart = LineChart()
    chart.title = "Loan Breakdown Over Time"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Year"

    data = Reference(ws, min_col=2, min_row=start_table_row, max_col=4, max_row=table_row - 1)
    categories = Reference(ws, min_col=1, min_row=start_table_row+1, max_row=table_row - 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, f"H{start_table_row}")

    return wb
